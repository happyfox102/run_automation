from openpyxl import load_workbook

wb = load_workbook("data.xlsx")
ws = wb.active

texts = []

for row in ws["A1:A10"]:  # диапазон из 10 ячеек
    for cell in row:
        texts.append(str(cell.value))

print(texts)
